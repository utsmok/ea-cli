"""
Easy Access Sheet Toolkit
September 2024
Samuel Mok / s.mok@utwente.nl / cip@utwente.nl
homepage: https://github.com/utsmok/easyaccesscli/

Note: only tested on windows systems

This script can be used to:
-> Read data from a CopyRight export
-> Process the data into a standard format for usage by people who check the documents on canvas
-> Export various sheets:
    -> per faculty
    -> all items
    -> only items that have been changed
    -> etc
-> Read back the data from all the sheets to produce a CopyRight 'import' sheet


Q U I C K    S T A R T
    Run with standard settings:
        > uv run easy_access_cli.py
    View cli instructions:
        > uv run easy_access_cli.py --help

I N S T A L L   U V ?
    UV is an all-in-one python manager.
    Install by opening Powershell (press windows key, type 'powershell', enter) and pasting the following lines:
        >
        >
    and press enter to install. For more info, see the uv docs: https://docs.astral.sh/uv/getting-started/installation/
    Once uv is installed, close PowerShell, start it again, and type
        > uv python install
    and the setup is all done! Now you can run the cli help with:
        > uv run easy_access_cli.py --help

This python file contains the following:
    - class EasyAccessToolkit with core functionality to ingest & process data from SURF's copyRight tool for easy access, and export various sheets for end-users
    - typer function cli provides a command line interface
    - helper classes File and Directory for handling... files and directories.
"""
import pathlib
import shutil

import openpyxl.worksheet
import openpyxl.worksheet.datavalidation
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles.differential import DifferentialStyle

import openpyxl.worksheet.table
import openpyxl.worksheet.worksheet
import typer
from typing_extensions import Annotated
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
import polars as pl
from pathlib import Path
import os
from datetime import datetime
import dotenv
from enum import Enum
import json
import openpyxl


print: callable = Console(emoji=True, markup=True).print
dotenv.load_dotenv('settings.env')

def info(text: str):
    """
    Prints an information message.
    """
    print(f"[cyan]:information: |> [/cyan] {text}")

def warn(text: str):
    """
    Prints a warning message.
    """
    print(f":warning: [bold red] Warning! [/bold red] |> {text}")

def cool(text: str):
    """
    Prints a nice message.
    """
    print(f":smiling_face_with_sunglasses: [yellow] {text} [/yellow]")



class Directory:
    """
    Simple class for directories + operations.
    Init with an absolute path, or a path relative to the current working directory.
    If the dir does not yet exist, it will be created. Disable this by setting the 'create_dir' parameter to False.
    """

    def __init__(self, path: str, create_dir: bool = True):
        self.input_path_str = path
        self.create_dir = create_dir

        # check if the path is absolute
        if pathlib.Path(path).is_absolute():
            self.full = pathlib.Path(path)
        else:
            self.full = pathlib.Path.cwd() / path

        self.post_init()

    def post_init(self) -> None:
        '''
        Checks to see if this is actually a dir,
        or create it if create_dir is set to True.
        '''
        if not self.full.exists():
            if self.create_dir:
                self.create()
            else:
                raise FileNotFoundError(f"Directory {self.full} does not exist and create_dir is set to False.")
        if not self.full.is_dir():
            raise NotADirectoryError(f"Directory {self.full} is not a directory.")


    @property
    def files(self) -> list['File']:
        '''
        Gets all files in the dir as a list of File objects.
        '''
        return [File(self.full / file) for file in self.full.iterdir() if file.is_file()]

    @property
    def files_r(self) -> list['File']:
        '''
        Recursively gets all files in the dir, so including files in subdirs, as a list of File objects.
        '''
        return [File(self.full / file) for file in self.full.rglob('*') if file.is_file()]

    @property
    def dirs(self, r: bool = False) -> list['Directory']:
        '''
        Returns a list of all dirs in this Directory as a list of Directory objects.
        If r is set to True, it will return all children dirs recursively.
        '''
        if not r:
            return [Directory(self, d) for d in self.full.iterdir() if d.is_dir()]
        if r:
            return [Directory(self, d) for d in self.full.rglob('*') if d.is_dir()]
    @property
    def newest_file(self) -> 'File':
        '''
        Returns the newest file in the dir as a File object.
        '''
        all_files = self.files
        return max(all_files, key=lambda x: x.created())

    @property
    def newest_file_r(self) -> str:
        '''
        Recursively gets the newest file in the dir, so including files in subdirs, as a File object.
        '''
        all_files = self.files_r
        return max(all_files, key=lambda x: x.created())

    @property
    def exists(self) -> bool:
        return self.full.exists()

    @property
    def is_dir(self) -> bool:
        return self.full.is_dir()

    def create(self) -> None:
        try:
            self.full.mkdir(parents=True, exist_ok=False)
        except FileExistsError:
            pass


    def __eq__(self, other) -> bool:
        return self.full == other.full

    def __str__(self):
        return str(self.full)

    def __repr__(self):
        return f"DirPath('{self.input_path_str}') -> {self.full}"

class File:
    '''
    Simple class for files + operations
    Parameters:
        path: str or Path
            relative from the current working directory.
            OR
            absolute path to the file.
            Should always end with the filename including extension.
    '''
    def __init__(self, path: str | pathlib.Path):
        self._path_init_str = str(path)

        assert isinstance(path, str) or isinstance(path, pathlib.Path)

        if isinstance(path, pathlib.Path):
            self._path = path
            self._name = path.name
            self._extension = path.suffix
            self._dir = Directory(str(self._path.absolute().parent))
        elif isinstance(path, str):
            if '/' in path:
                self._name = path.rsplit('/', 1)[-1]
                self._dir = Directory(path.rsplit('/', 1)[0], create_dir=True)
            else:
                self._name = path
                self._dir = Directory(os.getcwd())

            self._extension = self._name.split('.')[-1]
            self._path = self._dir.full / self._name

    @property
    def exists(self) -> bool:
        return self._path.exists()

    @property
    def is_file(self) -> bool:
        return self._path.is_file()

    @property
    def path(self) -> pathlib.Path:
        return self._path

    @property
    def name(self) -> str:
        return self._name

    @property
    def extension(self) -> str:
        return self._extension

    @property
    def dir(self) -> Directory:
        return self._dir

    @property
    def created(self) -> datetime:
        return datetime.fromtimestamp(self._path.stat().st_birthtime)

    @property
    def modified(self) -> datetime:
        return datetime.fromtimestamp(self._path.stat().st_mtime)

    def copy(self, new_path: str) -> 'File':
        shutil.copy(self._path, new_path)
        return File(new_path)

    def move(self, new_path: str) -> 'File':
        shutil.move(self._path, new_path)
        return File(new_path)

    def rename(self, new_name: str) -> 'File':
        self._path = self._dir.full / new_name
        return File(self._path)

    def __eq__(self, other) -> bool:
        return self._path == other.path

    def __str__(self):
        return str(self._path)

    def __repr__(self):
        if self._path_init_str != str(self._path):
            return f"FilePath('{self._path_init_str}') -> {self._path}"
        else:
            return f"FilePath('{self._path}')"

class Functions(str, Enum):
    """
    CLI option for picking which functions to run, see cli()
    """
    both = "both"
    read = "read"
    export = "export"

# The main CLI function:
def cli(do: Annotated[Functions, typer.Option(case_sensitive=False, help="Which tool to run: read in new data, export current data, or both.", rich_help_panel="Functions")] = Functions.read,
        changes: Annotated[bool, typer.Option(help="Only add items that have been changed to new faculty sheets.", rich_help_panel="Functions")] = True,
        other_sheet: Annotated[str | None, typer.Option(help="(relative) path to a xlsx sheet to read in instead of CopyRight Data.", rich_help_panel="Read in data from alternate source")] = None,
        copyright_export_dir: Annotated[str | None, typer.Argument(help="Overwrite location for files exported from CopyRight", rich_help_panel="Overwite Directory Locations")] = None,
        copyright_import_dir: Annotated[str | None, typer.Argument(help="Overwrite location for files to be imported back into CopyRight", rich_help_panel="Overwite Directory Locations")] = None,
        faculties_dir: Annotated[str | None, typer.Argument(help="Overwrite location for faculty sheets", rich_help_panel="Overwite Directory Locations")] = None,
        all_items_dir: Annotated[str | None, typer.Argument(help="Overwrite location for 'all items' sheet", rich_help_panel="Overwite Directory Locations")] = None,
        ):
    """
    Runs the Easy Access toolkit with the specified settings.\n
    Make sure that these two files are present in the current dir and contain the required info:\n\n
        'settings.env': The directories to use\n
        'department_mapping.json': The mapping between department names and faculty names\n
    \n
    Visit the repo for more instructions & the latest version: https://github.com/utsmok/ea-cli. (<- you can click this in your terminal!)\n
    \n\n
    Example usage\n
    --------------\n
    ea-cli\n
    ea-cli --do export\n
    ea-cli --no-changes\n
    ea-cli --do read --changes\n
    ea-cli --do both --copyright_export_dir 'C:/easy_access_sheets/cli_copyright_data'  \n
    """

    if do not in [Functions.both, Functions.read, Functions.export]:
        warn("No functions selected! Aborting the script. Next time, enable at least one of 'Function' options; for details run ea-cli --help.")
        cool("Thank you for using the Easy Access tool!")
        raise typer.Exit(code=1)

    if all(not x for x in [copyright_export_dir, copyright_import_dir, faculties_dir, all_items_dir]):
        dirs = None
    else:
        dirs={
            'copyright_export':copyright_export_dir,
            'copyright_import':copyright_import_dir,
            'faculties':faculties_dir,
            'all_items':all_items_dir,
        }

    tool = EasyAccessTool(functions=do, only_changes=changes, dirs=dirs, other_sheet=other_sheet)
    tool.run()

    cool("All done! Thank you for using the Easy Access tool!")

class EasyAccessTool:
    """
    This class contains all the actual functionality for handling the data and creating the sheets.
    """
    # which functions to run when self.run() is called
    settings: list[callable] = []

    # keep track of relevant files and directories
    files: dict[str,File]
    dirs: dict[str,Directory] = {
        'root':Directory(os.getcwd()),
        'copyright_export':Directory(os.getenv("COPYRIGHT_EXPORT_DIR")),
        'copyright_import':Directory(os.getenv("COPYRIGHT_IMPORT_DIR")),
        'all_items':Directory(os.getenv("ALL_ITEMS_DIR")),
        'faculties':Directory(os.getenv("FACULTIES_DIR")),
    }

    # the various dataframes created from / writing to .xlsx files
    raw_copyright_data: pl.DataFrame = pl.DataFrame() # data directly from copyRight
    copyright_data: pl.DataFrame = pl.DataFrame() # data with a bit of cleanup
    faculty_sheet_data: pl.DataFrame = pl.DataFrame() # data from the faculty sheets
    all_items_sheet_data: pl.DataFrame = pl.DataFrame() # data from the 'all_items' sheet

    # maps copyright data column 'departments' to faculties
    dept_mapping_path = File("department_mapping.json")
    DEPARTMENT_MAPPING = json.load(open(dept_mapping_path.path, encoding='utf-8'))

    # list of all found/used faculties
    faculties: list[str]

    # latest copyRight file & when it was created
    latest_file: File
    latest_file_date: str

    # starting excel style number for the data entry tables
    style_iter: int = 2

    # path to another sheet to read in instead of CopyRight data
    other_sheet: File | None = None

    # flag to indicate if there are no new items to add
    no_new_items: bool = False
    # standard column order for the complete data sheets
    column_order = ["material_id","period","department","course_code","course_name","url","filename","title","owner","filetype","classification","type","ml_prediction","manual_classification","manual_identifier","scope","remarks","auditor","last_change","status","google_search_file","isbn","doi","in_collection","pagecount","wordcount","picturecount","author","publisher","reliability","pages_x_students","count_students_registered","retrieved_from_copyright_on","workflow_status","faculty"]
    def __init__(self,
                    functions: Functions | None = Functions.both,
                    dirs: dict[str,str] | None = None,
                    only_changes: bool = True,
                    other_sheet: str | None = None
                ) -> None:
        """
        Parameters:
            setting:  str | None
                Pick which functions to run when self.run() is called. If no argument is passed, it will run all functions.
                pick from one of the presets below:
                    'none' -> don't run any functions
                    'all' -> run all functions
                    'new_data' -> read in new data, process it, create new faculty and all itemssheets
                    'read_sheets' -> read in faculty sheet data, process, create import sheet
            dirs: dict[str,str] | None
                A dict containing the str path to the directories to use. If not provided, it will use the default dirs from settings.env.
            only_changes: bool = True
                True (default): only add items that have been changed to the created sheets
                False: add all items from the CopyRight export to the created sheets
            other_sheets: list[str] | None
                A list of paths to additional .xlsx files to ingest instead the raw data from CopyRight.
        """

        # determine which functions to run
        # first check if we need to read in copyRight data (default), or other sheets

        if other_sheet:
            self.other_sheet = File(other_sheet)


        # export only new items, or all ingested items?
        self.only_changes = only_changes

        # set the functions to run

        if functions is None:
            self.settings = []
        elif functions == Functions.both:
            if not self.other_sheet:
                self.settings = [self.read_copyright_export, self.process_copyright_export,  # read in new data
                            self.read_all_items_sheet, self.read_faculty_sheets, # read in data manually added to sheets
                            self.create_import_sheet,  # from the old data, create a sheet to import into CopyRight
                            self.create_faculty_sheets, self.create_all_items_sheet] # create new sheets with new data
            else:
                self.settings = [self.read_other_sheet,  # read in new data
                                self.read_all_items_sheet, self.read_faculty_sheets, # read in data manually added to sheets
                                self.create_import_sheet,  # from the old data, create a sheet to import into CopyRight
                                self.create_faculty_sheets, self.create_all_items_sheet] # create new sheets with new data

        elif functions == Functions.read:
            if not self.other_sheet:
                self.settings = [self.read_copyright_export, self.process_copyright_export,  # read in new data
                            self.create_faculty_sheets, self.create_all_items_sheet] # create new sheets with new data
            else:
                self.settings = [self.read_other_sheet,self.process_copyright_export,  # read in new data
                    self.create_faculty_sheets, self.create_all_items_sheet # create new sheets with new data
                ]
        elif functions == Functions.export:
            self.settings = [self.read_faculty_sheets,  # read in data manually added to sheets
                            self.create_import_sheet,  # from the old data, create a sheet to import into CopyRight
                        ]
            if self.other_sheet:
                warn(f"Note: Only exporting data, so the contents other sheet {self.other_sheet} will have no effect on the output.")

        # if dirs is set, add them to the self.dirs dict
        if dirs:
            for key, value in dirs.items():
                if value:
                    self.dirs[key] = Directory(value)

    def run(self) -> None:
        """
        Runs the functions as specified in the settings dict.
        """
        for func in self.settings:
            func()

    def read_other_sheet(self) -> None:
        """
        Reads in the data from another sheet as the datasource, instead of using CopyRight data.
        Sheet should be formatted in the same way as the faculty output sheets.
        It will read in the first sheet in the .xlsx file.
        It will do a quick check on the columns in the sheets.
        """

        info(f"Reading in data from {self.other_sheet.name}")
        self.copyright_data = pl.read_excel(self.other_sheet.path)
        self.latest_file_date = self.other_sheet.modified.strftime("%Y-%m-%d")
        info(f"Read {len(self.copyright_data)} items from {self.other_sheet.name}. Item was lasted changed on {self.latest_file_date}")

        if 'workflow_status' not in self.copyright_data.columns:
                self.copyright_data = self.copyright_data.with_columns(
                    pl.Series("workflow_status", ["ToDo"] * len(self.copyright_data))
                )
        if  'retrieved_from_copyright_on' not in self.copyright_data.columns:
            if 'added_to_sheet_on' not in self.copyright_data.columns:
                self.copyright_data = self.copyright_data.with_columns(
                    pl.Series("retrieved_from_copyright_on", [self.latest_file_date] * len(self.copyright_data))
                )
            else:
                self.copyright_data = self.copyright_data.rename(
                    {"added_to_sheet_on":'retrieved_from_copyright_on'}
                )

        self.latest_file_date = max(self.copyright_data.select(pl.col("retrieved_from_copyright_on")).to_series().to_list())
        self.copyright_data = self.copyright_data.select(self.column_order)


    def read_copyright_export(self) -> None:
        info(f"Reading in newest Copyright Data from directory: {self.dirs['copyright_export']}")
        try:
            all_files = self.dirs['copyright_export'].files
            self.latest_file = max(all_files, key=lambda x: x.created)
            self.latest_file_date = self.latest_file.created.strftime("%Y-%m-%d")
            info(f"Selected newest copyright export file:\n          {self.latest_file.name}\n          created @ {self.latest_file_date}")
            self.raw_copyright_data = pl.read_excel(self.latest_file.path)
        except FileNotFoundError:
            warn(f"No files found in {self.dirs['copyright_export']}")
            raise typer.Exit(code=1)
        except PermissionError:
            warn(f"Permission denied to read {self.latest_file.name}")
            raise typer.Exit(code=1)
        except ValueError:
            warn("No files found in {self.dirs['copyright_export']}")
            raise typer.Exit(code=1)

    def process_copyright_export(self) -> None:
        """
        Process the raw copyright data:
        rename column headers, add extra columns, format some data, and match to faculty.

        If 'only_changes' it will compare this data to the items present in the faculty sheets,
        and only include new items in the export.
        """
        if self.copyright_data.is_empty():
            self.copyright_data = self.raw_copyright_data.rename(
                lambda col: col.replace(" ", "_")
                .replace("#", "count_")
                .replace("*", "x")
                .lower()
                ).with_columns(
                pl.Series("retrieved_from_copyright_on", [self.latest_file_date] * len(self.raw_copyright_data)),
                pl.Series("workflow_status", ["ToDo"] * len(self.raw_copyright_data)),
                pl.col("last_change").dt.strftime("%Y-%m-%d"),
                faculty=pl.col("department").replace_strict(
                    self.DEPARTMENT_MAPPING, default="Unmapped"
                ),
            )

        if self.only_changes:
            self.read_faculty_sheets()
            if self.faculty_sheet_data.is_empty():
                info("No faculty sheets found. Adding all items without checking for changes.")
            else:
                # compare self.copyright_data with self.faculty_sheet_data
                # only keep items from self.copyright_data with a value in col material_id that is not in self.faculty_sheet_data
                # AND items with a matching material_id but a different value in col last_change
                not_in_faculty = self.copyright_data.join(
                    self.faculty_sheet_data,
                    on="material_id",
                    how="anti"
                )
                matching_id_diff_change = self.copyright_data.join(
                    self.faculty_sheet_data,
                    on="material_id",
                    how="inner"
                ).filter(
                    pl.col("last_change") != pl.col("last_change_right")
                ).select(
                    pl.all().exclude("last_change_right")
                ).drop_nulls(
                    pl.col("material_id")
                ).filter(
                    pl.col("status") == "Deleted"
                )

                if not_in_faculty.is_empty():
                    if matching_id_diff_change.is_empty():
                        info("No new items to add!")
                        self.no_new_items = True
                    else:
                        self.copyright_data = matching_id_diff_change
                if not matching_id_diff_change.is_empty():
                    self.copyright_data = pl.concat([not_in_faculty, matching_id_diff_change])
                else:
                    self.copyright_data = not_in_faculty


        self.faculties = self.copyright_data.select(pl.col("faculty").unique()).to_series().to_list()

    def create_faculty_sheets(self) -> None:
        """
        Splits the processed copyright data into one sheet per faculty
        and exports the result to excel sheets.
        """
        if self.faculties:
            self.faculties.sort()
        for faculty in self.faculties:
            faculty_dir = Directory(self.dirs['faculties'].full / faculty)
            if faculty is None or faculty == "":
                faculty = "no_faculty_found"
            filename = f"{faculty}_{self.latest_file_date}.xlsx"
            i = 1
            while os.path.exists(faculty_dir.full / filename):
                filename = f"{faculty}_{self.latest_file_date}_{i}.xlsx"
                i += 1

            faculty_data = self.copyright_data.filter(pl.col("faculty") == faculty)

            if faculty_data.is_empty():
                warn(f"No items found for faculty {faculty}.")

            faculty_data.write_excel(faculty_dir.full / filename)
            info(f"Created sheet {filename}")
            self.finalize_sheet(File(str(faculty_dir.full / filename)))


    def finalize_sheet(self, file: File) -> None:
        """
        Add the data entry sheet to a fresh Faculty Excel file.
        This sheet will contain a selection of columns, will be styled,
        and contain dropdowns for data entry.
        """

        wb = openpyxl.load_workbook(filename = str(file.path))
        wb.active.title = 'Complete data'

        # Create the Data Entry sheet
        # -----------------------------

        entry_sheet: openpyxl.worksheet.worksheet.Worksheet = wb.create_sheet('Data entry', index=1)
        keep_cols = [6, 34, 14, 16, 17, 13, 1, 8, 9, 28, 3, 5, ""]
        col_names = ['url', 'workflow_status', 'manual_classification', 'scope', 'remarks', 'ml_prediction',
                'material_id', 'title', 'owner', 'author', 'department', 'course_name']
        max_row = 0
        url = False

        for new_col, old in enumerate(keep_cols, start=1):
            if isinstance(old, str):
                break
            for row, cell in enumerate(wb.active.iter_rows(min_col=old, max_col=old, values_only=True), start=1):
                if row == 1 and cell[0]=='url':
                    url = True
                elif row == 1:
                    url = False
                if not url:
                    entry_sheet.cell(row=row, column=new_col).value = cell[0]
                else:
                    entry_sheet.cell(row=row, column=new_col).hyperlink = cell[0]
                if row > max_row:
                    max_row = row

        for col, name in enumerate(col_names, start=1):
            entry_sheet.cell(row=1, column=col, value=name)

        # Dropdown items for certain cells
        # -----------------------------------
        dropdowndata = [(2,'B', '"ToDo,Done,InProgress"'), # workflow status
                        (3,'C', '"open access, eigen materiaal - powerpoint, eigen materiaal - overig, lange overname, eigen materiaal - titelindicatie"'), # manual classification
                        ]
        for colnum, col_letter, itemlist in dropdowndata:
            dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1=itemlist, allow_blank=False)
            dv.error = "Please select a valid option from the list"
            dv.errorTitle = "Invalid option"
            dv.prompt= "Please select from the list"
            dv.promptTitle = "List selection"
            entry_sheet.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{max_row}")

        # Style as table
        # -----------------
        table = ExcelTable(displayName="DataEntry", ref=f"A1:L{max_row}")
        tabstyle = TableStyleInfo(
            name=f'TableStyleMedium{self.style_iter}',
            showRowStripes=True,
        )
        self.style_iter = self.style_iter+1
        table.tableStyleInfo = tabstyle
        entry_sheet.add_table(table)

        wb.save(filename = str(file.path))

    def create_all_items_sheet(self) -> None:

        """
        Add all items in the current Copyright data to a single sheet for CIP ease of use.
        """
        if not self.no_new_items:
            filename = f"all_items_{self.latest_file_date}.xlsx"
            i = 1
            while os.path.exists(self.dirs['all_items'].full / filename):
                filename = f"all_items_{self.latest_file_date}_{i}.xlsx"
                i += 1

            self.copyright_data.write_excel(self.dirs['all_items'].full / filename)
            info(f"Created sheet: {self.dirs['all_items'].full / filename}")
    def read_faculty_sheets(self) -> None:
        """
        Reads in all data from all sheets in the faculties dir
        and stores it in self.faculty_sheet_data as a single concatted dataframe.
        """
        self.faculty_sheet_data = self.read_sheets(self.dirs['faculties'].files_r)

    def read_all_items_sheet(self) -> None:
        """
        Reads in all data from all 'all_items' sheets
        and stores it in self.all_items_sheet_data as a single concatted dataframe.
        """
        if False:
            self.all_items_sheet_data  = self.read_sheets(self.dirs['all_items'].files_r)
        warn("read_all_items_sheet isn't used at the moment. If you want to import data from an 'all_items' sheet, please use the --other-sheet option in the cli instead.")

    def read_sheets(self, files: list[File]) -> pl.DataFrame:
        """
        Reads the data from all Excel files in the list of files.
        Currently only reads in the data from the sheet 'Complete data'.
        TODO: Handle ingestion of data the faculty added to the sheet.
        """
        file_data = []
        for file in files:
            if file.extension not in ['.xls', '.xlsx']:
                warn(f'{file.name} is not an excel file, skipping.')
                continue
            try:
                current_data = pl.read_excel(file.path, sheet_name='Complete data')
            except Exception as e:
                current_data = pl.read_excel(file.path)

            current_data = self.validate_ea_sheet(current_data, file)
            if current_data.is_empty():
                continue
            else:
                file_data.append(current_data)
        if file_data:
            result: pl.DataFrame = pl.concat(file_data)
        else:
            result = pl.DataFrame()
        return result.unique()

    def validate_ea_sheet(self, df: pl.DataFrame, file:File) -> pl.DataFrame:
        """
        For a given dataframe created from an EA excel sheet,
        check the data for errors.
        If found, try to fix, else print the errors.
        If the sheet is not validated, return an empty dataframe.

        Current implementation is bare:
        - is sheet empty? if yes: print error
        - set all columns to type str

        TODO: Implement this function fully.
        TODO: handle multiple sheets in the same file

        """
        valid = True
        errlist = []
        if df.is_empty():
            valid = False
            errlist.append("Sheet is empty")
        if valid:
            # set all columns to type str
            df = df.with_columns(pl.exclude(pl.Utf8).cast(str))
            # check that the sheet has the correct columns
            ...
        if valid:
            # check the values in the columns
            ...
        if not valid:
            info(f"Errors in sheet {file}:")
            for err in errlist:
                warn(err)
            return pl.DataFrame()
        else:
            return df

    def create_import_sheet(self) -> None:
        '''
        combine self.faculty_sheet_data and self.all_items_sheet_data
        clean it up
        change from UT Easy Access format to SURF CopyRight format
        create & export an .xlsx sheet that can be sent to SURF to be imported into CopyRight.
        '''
        #TODO
        ...

if __name__ == "__main__":
    typer.run(cli)
